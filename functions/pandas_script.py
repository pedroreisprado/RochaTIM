import pandas as pd
import xlrd
import openpyxl
from openpyxl import load_workbook
from PyPDF2 import PdfReader
# from config import PATH_CONTAS

import re
import zipfile
import os

# from config import CNPJ1,CNPJ2,CNPJ3,CNPJ4,CNPJ5,CNPJ6
# from functions.logger import logger

# def getTM(desc,estoque,finan):
#     aux_files = 'C:\\2RFP\\jm\\aux_files\\depara.xls'
#     df = pd.read_excel(aux_files)
#     df_filter = df[(df['Tipo'] == desc) & (df['Estoque'] == estoque) & (df['Finan'] == finan)]
#     if not df_filter.empty:
#      return(df_filter.iloc[0,1])
#     return 0

def readClient():
    base_cliente = 'C:\\Users\\Firework\Desktop\\2RFP_SCRIPTS\\Rocha\\aux_files\\clientes.xlsx'
    df = pd.read_excel(base_cliente)
    cliente = df['GSM'].tolist()
    return cliente

PATH_CONTAS = 'C:\\Users\\Firework\\Desktop\\2RFP_SCRIPTS\\Rocha\\pdf\\'


def extract_info_from_pdf(numero_cliente):
    text = ""
    with open(f'{PATH_CONTAS}{numero_cliente}.pdf', "rb") as pdf_file:
        pdf_reader = PdfReader(pdf_file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text += page.extract_text()

    # Define os padrões de expressões regulares para extrair informações específicas
    patterns = {
        'nome': r'(?<=Olá, )\b[A-Z\s]+\b',
        'endereço': r'Endereço:\s*([^\n]*)',
        'total': r'Total:  R\$ ([0-9,.]+)',
        'vencimento': r'Vencimento:  (\d{2}/\d{2}/\d{4})',
        'fatura': r'Fatura:  (\d+)',
        'periodo': r'Periodo:  (\d{2}/\d{2} a \d{2}/\d{2})',
        'emissao': r'Emissão:  (\d{2}/\d{2}/\d{4})',
        'numero_tim': r'Numero Tim:  (\d{2} \d{5}-\d{4})',
        'cliente': r'Cliente:  ([0-9]+)',
        'acessos': r'Acessos: (\d+)',
        'identificador_debito': r'Identificador do Debito automatico:  (\d+)',
        'referencia': r'Referencia:  (\w+/\d{4})',
        'cod_barras': r'Codigo de barras:  (\d+ - \d+ \d+ - \d+ \d+)',
        'numeros_telefone': r'Numero \d: (\d{2} \d{5}-\d{4})'  # Atualizado para capturar diretamente os números de telefone
    }
    
    # Inicializa um dicionário para armazenar as informações extraídas
    extracted_info = {}
    
    # Itera sobre os padrões e extrai as informações
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            extracted_info[key] = match.group(1).strip() if match.group(1) else ""  # Verifica se há correspondência antes de acessar o grupo
    
    # Se houver informações sobre os números de telefone, adiciona ao dicionário
    if 'acessos' in extracted_info:
        num_acessos = int(extracted_info['acessos'])
        for i in range(1, num_acessos + 1):
            key = f'numero_{i}'
            pattern = patterns['numeros_telefone']  # Não precisamos mais substituir o número na expressão regular
            match = re.search(pattern, text)
            if match:
                extracted_info[key] = match.group(1).strip() if match.group(1) else ""  # Verifica se há correspondência antes de acessar o grupo
    
    return extracted_info

def extract_text_from_pdf(numero_cliente,nome,cpf):
    cpf = str(cpf)
    numero_cliente = str(numero_cliente)
    nome = str(nome)
    text = ""
    with open(f'{PATH_CONTAS}{numero_cliente}.pdf', "rb") as pdf_file:
        pdf_reader = PdfReader(pdf_file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text += page.extract_text()
            # Divida o texto em linhas
    lines = text.split('\n')
    # Escolha a linha que você deseja imprimir
    linha_especifica = 9  # Substitua pelo índice da linha que você deseja imprimir

    # Verifique se o índice é válido e imprima a linha específica
    rua = (lines[6])
    bairro = (lines[7])
    cidade_estado = (lines[8])
    ser_chamado = (lines[15]).replace("Olá,", "").replace("!", "")
    # Encontrar o numero identificacao usando a expressao regular
    match = re.search(r'Nº de Identificação do documento: (\d+)', text)
    # Verificar se houve correspondencia e extrair o numero identificacao
    if match:
        fatura = match.group(1)

    for i,line in enumerate(lines):
        if 'Escaneie o QR Code ao lado e' in line:
            codigo_debt = (lines[i-1])
            break
    
    for i,line in enumerate(lines):
        if 'CFOP' in line:
            referencia = (lines[i+1])
            break
    for i,line in enumerate(lines):
        if 'EMISSÃO' in line and 'VENCIMENTO' not in line:
            emissao = (lines[i+1])
            break
    for i,line in enumerate(lines):
        if 'POSTAGEM' in line:
            postagem = (lines[i+1]).replace('EMISSÃO',"")
            break
    for i,line in enumerate(lines):
        if 'Cliente' in line and 'VENCIMENTO' not in line:
            cliente = (lines[i]).replace('Cliente ',"")
            break
    for i,line in enumerate(lines):
        if 'acessos' in line and 'VENCIMENTO' not in line:
            acessos = (lines[i]).replace('Quantidade de acessos ',"")
            break
    for i,line in enumerate(lines):
        if f'{referencia}' in line and 'fatura' not in line:
            periodo = (lines[i+1])
            break
    for i,line in enumerate(lines):
        if f'{referencia}' in line and 'fatura' in line:
            codigo_barras = (lines[i][16:79])
            break

    for i,line in enumerate(lines):
        if 'Pague com Pix' in line and 'VALOR' in line:
            valor_parcela = (lines[i+1]).replace('Pague com Pix ',"").replace('R$ ',"")
            break
    for i,line in enumerate(lines):
        if 'VENCIMENTO' in line and 'VALOR' not in line:
            vencimento = (lines[i+1])
            break

    for i,line in enumerate(lines):
        if 'plano oferece' in line:
            teste = (lines[i])
            posicao_inicio = (lines[i]).find("Vantagens que seu plano oferece:") + len("Vantagens que seu plano oferece:")
            # Extrair os próximos 10 caracteres após a posição de "Vantagens que seu plano oferece:"
            numero_1 = (lines[i][posicao_inicio:posicao_inicio + 30])

    linhas_encontradas = 0
    acessos_int = int(acessos)
    # Inicializar um dicionário para armazenar os números de telefone
    numeros_telefone = {}

    # Iterar sobre as linhas para encontrar todos os números
    for i, line in enumerate(lines):
        if 'plano oferece' in line:
            # Encontrar a posição de "Vantagens que seu plano oferece:"
            posicao_inicio = line.find("Vantagens que seu plano oferece:") + len("Vantagens que seu plano oferece:")
            # Extrair os próximos 30 caracteres após a posição de "Vantagens que seu plano oferece:"
            numero = line[posicao_inicio:posicao_inicio + 30]

            # Armazenar o número de telefone no dicionário
            numeros_telefone[f'numero_{i+1}'] = numero

    # Imprimir os números de telefone armazenados no dicionário
    atual = 1
    for chave, valor in numeros_telefone.items():
        # print(f"{valor}")
        if atual == 1:
            numero_1 = valor
            atual += 1  # Incrementar o valor de atual
            continue
        if atual == 2:
            numero_2 = valor
            atual = 3  # Atribuir o valor 3 a atual
            continue
        if atual == 3:
            numero_3 = valor
            atual = 4  # Atribuir o valor 4 a atual
            continue
        if atual == 4:
            numero_4 = valor
            atual = 5  # Atribuir o valor 5 a atual
            continue
            
    print(f'valor:{valor_parcela}')

    # Carregar o arquivo Excel
    workbook = load_workbook('C:\\Users\\Firework\\Desktop\\2RFP_SCRIPTS\\Rocha\\excel\\devolutivas.xlsx')

    # Selecionar a planilha desejada
    sheet = workbook['devolutivas']

    # Determinar a última linha preenchida na coluna A
    ultima_linha = sheet.max_row

    # Adicionar as variáveis abaixo das linhas existentes
    sheet.cell(row=ultima_linha + 1, column=1).value = numero_cliente
    sheet.cell(row=ultima_linha + 1, column=2).value = nome
    sheet.cell(row=ultima_linha + 1, column=3).value = ser_chamado
    sheet.cell(row=ultima_linha + 1, column=4).value = fatura
    sheet.cell(row=ultima_linha + 1, column=5).value = periodo
    sheet.cell(row=ultima_linha + 1, column=6).value = emissao
    sheet.cell(row=ultima_linha + 1, column=7).value = postagem
    sheet.cell(row=ultima_linha + 1, column=8).value = vencimento
    sheet.cell(row=ultima_linha + 1, column=9).value = valor_parcela
    sheet.cell(row=ultima_linha + 1, column=10).value = cliente
    sheet.cell(row=ultima_linha + 1, column=11).value = cpf
    sheet.cell(row=ultima_linha + 1, column=12).value = acessos
    sheet.cell(row=ultima_linha + 1, column=13).value = rua
    sheet.cell(row=ultima_linha + 1, column=14).value = bairro
    sheet.cell(row=ultima_linha + 1, column=15).value = cidade_estado
    sheet.cell(row=ultima_linha + 1, column=16).value = codigo_debt
    sheet.cell(row=ultima_linha + 1, column=17).value = referencia
    sheet.cell(row=ultima_linha + 1, column=18).value = codigo_barras
    sheet.cell(row=ultima_linha + 1, column=19).value = numero_1
    if acessos_int == 4:
        sheet.cell(row=ultima_linha + 1, column=20).value = numero_2
        sheet.cell(row=ultima_linha + 1, column=21).value = numero_3
        sheet.cell(row=ultima_linha + 1, column=22).value = numero_4
    if acessos_int == 3:
        sheet.cell(row=ultima_linha + 1, column=20).value = numero_2
        sheet.cell(row=ultima_linha + 1, column=21).value = numero_3
    if acessos_int == 2:
        sheet.cell(row=ultima_linha + 1, column=20).value = numero_2


    # Salvar as alterações
    workbook.save('C:\\Users\\Firework\\Desktop\\2RFP_SCRIPTS\\Rocha\\excel\\devolutivas.xlsx')




# def finalFiles():
#     folder_path = ('C:\\2RFP\\jm\\functions\\result')
#     arquivos = []
#     for filename in os.listdir(folder_path):
#         if filename.endswith('.txt'):
#             base_name = os.path.splitext(filename)[0]
#             arquivos.append(base_name)

#     return arquivos

# def finalResult(item):
#     folder_path = ('C:\\2RFP\\jm\\functions\\result')
#     file_path = f'{folder_path}\\{item}.txt'
#     with open(file_path, 'r+') as file:
#         content = file.read()
#         return content

# def putResult(item,status):
#     result_file_path = os.path.join(f'C:\\2RFP\\jm\\functions\\result\\{item}.txt')
#     if os.path.exists(result_file_path):
#         with open(result_file_path, 'r+') as file:
#             # Lê o conteúdo do arquivo
#             content = file.read()
#             if content == status:
#                 return  
#     with open(result_file_path, 'w') as file:
#         file.write(status)
#         return 
    
# def orderStatus(item,status):
#     status_file_path = os.path.join(f'C:\\2RFP\\jm\\functions\\status\\{item}.txt')
#     if os.path.exists(status_file_path):
#         with open(status_file_path, 'r+') as file:
#             # Lê o conteúdo do arquivo
#             content = file.read()
#             if content >= status:
#                 return 'skip'
#             return 'exec'
#     with open(status_file_path, 'w') as file:
#         file.write('1')
#         return 'new'


# def updateStatus(item,status):
#     status_file_path = os.path.join(f'C:\\2RFP\\jm\\functions\\status\\{item}.txt')
#     if os.path.exists(status_file_path):
#         with open(status_file_path, 'r+') as file:
#             file.seek(0)
#             file.write(status)
#             file.truncate() 
#             return

# def extract_text_from_pdf(pdf_file_path):
#     text = ""
#     with open(pdf_file_path, "rb") as pdf_file:
#         pdf_reader = PdfReader(pdf_file)
#         for page_num in range(len(pdf_reader.pages)):
#             page = pdf_reader.pages[page_num]
#             text += page.extract_text()
#     return text

# def search_value_in_pdf(pdf_file_path, value_to_search):
#     found = False
#     extracted_text = extract_text_from_pdf(pdf_file_path)
#     if value_to_search in extracted_text:
#         found = True
#     return found

# def getNumber(file):

#     file_path = 'C:\\2RFP\\jm\\data\\' + file

#     df = pd.read_excel(file_path)

#     values_column_1 = df.iloc[:, 0].tolist()

#     return values_column_1


# def clearBlankSpace(item,action):
#     arquivo_origem = f'C:\\2RFP\\jm\\info_ERP\\{item}.txt'
#     arquivo_saida = f'C:\\2RFP\\jm\\info_ERP\\{item}.txt'
#     if 'item' in action:
#         arquivo_origem = f'C:\\2RFP\\jm\\info_ERP\\item\\{item}.txt'
#         arquivo_saida = f'C:\\2RFP\\jm\\info_ERP\\item\\{item}.txt'
#     # Ler o arquivo e remover espaços em branco
#     with open(arquivo_origem, 'r') as arquivo:
#         linhas_sem_espacos = [linha.strip() for linha in arquivo if linha.strip()]

#     # Salvar as linhas sem espaços em branco no arquivo de saída
#     with open(arquivo_saida, 'w') as arquivo_saida:
#         for linha in linhas_sem_espacos:
#             arquivo_saida.write(f"{linha}\n")
#     return True

# def firstApprove(item):
#     status = '1'
#     status_now = orderStatus(item,status)
#     if 'exec' in status_now:
#         logger(f'Inicio de check {item}')
#         caminho_arquivo = f'C:\\2RFP\\jm\\info_nf\\{item}.txt'
#         df = pd.read_csv(caminho_arquivo, sep=',', header=None, encoding= 'latin1')
#         tipo = df.iloc[0, 0]

#         if df.iloc[0,6] <= 0:
#             error = "nao foi possivel localizar o numero do pedido"
#             putResult(item,error)
#             status = '99'
#             updateStatus(item,status) 
#             return 
#         if 'CONSUMO' not in tipo:
#             return "Pedido é diferente de consumo"
        
#         CNPJ = [CNPJ1,CNPJ2,CNPJ3,CNPJ4,CNPJ5,CNPJ6]
#         nf = df.iloc[0, 3]
#         nf = f"{nf:06d}"
#         nf = f"{nf[:3]}.{nf[3:]}"

#         pdf_file_path = f"C:\\2RFP\\jm\\pdf\\{item}.pdf"
#         found_nf = search_value_in_pdf(pdf_file_path,nf)
#         logger(f'Verificação do numero da NF')
#         if not found_nf:  
#             error = "Numero infomado da NF nao corresponde na NF"
#             putResult(item,error)
#             status = '99'
#             updateStatus(item,status) 
#             return 
#         verif_cnpj = None
#         for item in CNPJ:
#             found_cnpj = search_value_in_pdf(pdf_file_path,item)
#             logger(f'Verificação do CNPJ')
#             if found_cnpj:
#                 verif_cnpj = True
#         if verif_cnpj is None:
#             error = "CNPJ informado nao corresponde na NF"
#             putResult(item,error)
#             status = '99'
#             updateStatus(item,status) 
#             return 
#         updateStatus(item,status)   
    

# def getInfoERP(item):
#     logger(f'Informações {item}')
#     caminho_arquivo = f'C:\\2RFP\\jm\\info_nf\\{item}.txt'
#     df = pd.read_csv(caminho_arquivo, sep=',', header=None, encoding= 'latin1')
#     tipo = df.iloc[0, 0]
#     financeiro = df.iloc[0,1]
#     estoque = df.iloc[0,2]
#     order = df.iloc[0,6]
#     filial = df.iloc[0,7]
#     empresa = df.iloc[0,8]
#     nf = df.iloc[0,3]
#     return financeiro, estoque, order, filial, empresa, tipo, nf

# def erpAprrove(item, valor, ncm):
#     pdf_file_path = f"C:\\2RFP\\jm\\pdf\\{item}.pdf"
#     # Lista para armazenar NCMs não encontrados na NF
#     not_found_ncm = []
    
#     for ncm_value in ncm:
#         found_ncm = search_value_in_pdf(pdf_file_path, ncm_value)
#         logger(f'Verificando NCM {ncm_value}')
#         if not found_ncm:
#             log = f'NCM {ncm_value} nao localizado na NF'
#             return log
        
#     verif_valor = search_value_in_pdf(pdf_file_path, valor)
#     if not verif_valor:
#         logger(f'O valor de R$ {valor}, nao corresponde com a NF')
#         log = f'O valor de R$ {valor}, nao corresponde com a NF'
#         return log
        
#     return "ok"
        
# def getInTXT(item,action):
#     caminho_arquivo = f'C:\\2RFP\\jm\\info_ERP\\{item}.txt'
#     coll = 'NCM'
#     if 'item' in action:
#         caminho_arquivo = f'C:\\2RFP\\jm\\info_ERP\\item\\{item}.txt'
#         coll = 'Produto'
    
#     # Especifica que o dtype da coluna NCM deve ser string ao ler o arquivo CSV
#     df = pd.read_csv(caminho_arquivo, dtype={coll: str})
#     ncm_values = df[coll].tolist()
#     return ncm_values



            



